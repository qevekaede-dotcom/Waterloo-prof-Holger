#!/usr/bin/env python3
"""Build the Excel workbook attached to Roy's project update."""

from pathlib import Path
import csv
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "thermo_candidates" / "SrCu2SnS4"
OUTPUT = ROOT / "outputs" / "roy_email" / "SrCu2SnS4_results.xlsx"
ATTACHMENT = ROOT / "submission_to_roy" / "READY_TO_ATTACH" / OUTPUT.name

NAVY = "17324D"
TEAL = "147D80"
RED = "B64B4B"
LIGHT_TEAL = "EAF3F3"
LIGHT_BLUE = "EEF3F7"
LIGHT_GRAY = "F3F5F7"
MID_GRAY = "C9D2DA"
TEXT = "24313D"
WHITE = "FFFFFF"

thin_gray = Side(style="thin", color=MID_GRAY)


def read_csv(path):
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.font = Font(name="Calibri", size=17, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[cell.row].height = 30


def style_section(cell, text):
    cell.value = text
    cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    cell.fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    cell.alignment = Alignment(vertical="center")


def style_header(row_range):
    for row in row_range:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color=TEAL))


def set_number_format(ws, cell_range, number_format):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = number_format


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def setup_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_notes(wb):
    ws = wb.active
    ws.title = "Notes"
    setup_sheet(ws)
    style_title(ws, "A1:H1", "My first SrCu2SnS4 calculation")

    ws.merge_cells("A3:H4")
    ws["A3"] = (
        "I started with the experimentally observed Materials Project structure mp-16988. "
        "I first checked the plane-wave cutoff and k-point mesh, then relaxed the structure, "
        "ran SCF and dense NSCF calculations, and used the eigenvalues in BoltzTraP2."
    )
    ws["A3"].font = Font(name="Calibri", size=10, color=TEXT)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    style_section(ws["A6"], "What I used")
    ws.merge_cells("A6:H6")
    settings = [
        ("Structure", "Materials Project mp-16988; experimentally observed; ICSD-356"),
        ("DFT setup", "PBE, scalar-relativistic pseudopotentials; SOC not included"),
        ("Cutoffs", "ecutwfc = 90 Ry; ecutrho = 720 Ry"),
        ("Relaxation", "vc-relax with 4x4x2 k-point mesh"),
        ("Final SCF", "5x5x3 k-point mesh"),
        ("Dense NSCF", "12x12x6 mesh; 100 irreducible k points; 140 bands"),
        ("BoltzTraP2", "300-900 K; n-type and p-type; 1e19 to 1e21 cm^-3"),
    ]
    ws["A7"] = "Item"
    ws["B7"] = "My setting / note"
    style_header(ws["A7:B7"])
    for idx, (item, detail) in enumerate(settings, start=8):
        ws.cell(idx, 1, item).font = Font(bold=True, color=TEXT)
        ws.cell(idx, 2, detail)
        ws.cell(idx, 2).alignment = Alignment(wrap_text=True)
    add_table(ws, f"A7:B{7 + len(settings)}", "SettingsTable")

    style_section(ws["A17"], "What I found")
    ws.merge_cells("A17:H17")
    findings = [
        ("Relaxed symmetry", "P3_121 (No. 152) from symmetry analysis"),
        ("Cell volume", "543.884 to 546.507 A^3 (+0.48%)"),
        ("PBE gap estimate", "0.3445 eV indirect gap on the dense k-point grid"),
        ("Transport trend", "Under the same constant-tau assumption, sampled p-type PF/tau is generally larger"),
        ("Example", "At 300 K: p-type, about 1e20 cm^-3, S = 157.2 uV/K"),
    ]
    ws["A18"] = "Result"
    ws["B18"] = "My observation"
    style_header(ws["A18:B18"])
    for idx, (item, detail) in enumerate(findings, start=19):
        ws.cell(idx, 1, item).font = Font(bold=True, color=TEXT)
        ws.cell(idx, 2, detail)
        ws.cell(idx, 2).alignment = Alignment(wrap_text=True)
    add_table(ws, f"A18:B{18 + len(findings)}", "FindingsTable")

    style_section(ws["A25"], "What is not finished")
    ws.merge_cells("A25:H25")
    ws.merge_cells("A26:H28")
    ws["A26"] = (
        "BoltzTraP2 gives sigma/tau and electronic kappa/tau. I have not reported the "
        "electronic-only value as the final zT. Absolute conductivity still needs a "
        "relaxation time, and full zT also needs lattice thermal conductivity. The p-type "
        "comparison is conditional on using the same constant-tau assumption for both carrier types."
    )
    ws["A26"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A26"].fill = PatternFill("solid", fgColor="FFF4E5")
    ws["A26"].border = Border(left=Side(style="medium", color="D8892B"))

    ws["A30"] = "Structure source"
    ws["B30"] = "https://materialsproject.org/materials/mp-16988"
    ws["B30"].hyperlink = ws["B30"].value
    ws["B30"].style = "Hyperlink"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 78
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 3
    ws.freeze_panes = "A7"


def build_transport_data(wb, full_rows):
    ws = wb.create_sheet("Transport_Data")
    setup_sheet(ws)
    style_title(ws, "A1:M1", "BoltzTraP2 processed transport data")
    ws.merge_cells("A2:M3")
    ws["A2"] = (
        "Scalar trace output under the constant-relaxation-time approximation. "
        "PF/tau is calculated in this workbook as (S x 1e-6)^2 x sigma/tau. "
        "No electronic-only zT column is included because it is not the full zT."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    headers = [
        "Temperature (K)",
        "Carrier type",
        "Carrier density (cm^-3)",
        "Signed density (cm^-3)",
        "Chemical potential (Ry)",
        "Seebeck S (uV/K)",
        "sigma/tau (ohm^-1 m^-1 s^-1)",
        "kappa_e/tau (W m^-1 K^-1 s^-1)",
        "PF/tau (W m^-1 K^-2 s^-1)",
        "Hall coefficient (m^3/C)",
        "Smoothed DOS (Ha^-1 uc^-1)",
        "Cv (J mol^-1 K^-1)",
        "chi (m^3 mol^-1)",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(5, col, header)
    style_header(ws["A5:M5"])

    for excel_row, row in enumerate(full_rows, start=6):
        values = [
            int(float(row["temperature_K"])),
            row["carrier_type"],
            float(row["carrier_density_cm-3"]),
            float(row["signed_density_cm-3"]),
            float(row["chemical_potential_Ry"]),
            float(row["seebeck_uV_K"]),
            float(row["sigma_over_tau_ohm-1_m-1_s-1"]),
            float(row["kappa_e_over_tau_W_m-1_K-1_s-1"]),
            None,
            float(row["hall_m3_C"]),
            float(row["smoothed_DOS_Ha-1_uc-1"]),
            float(row["cv_J_mol-1_K-1"]),
            float(row["chi_m3_mol-1"]),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(excel_row, col, value)
        ws.cell(excel_row, 9, f"=(F{excel_row}/1000000)^2*G{excel_row}")

    end = 5 + len(full_rows)
    add_table(ws, f"A5:M{end}", "TransportDataTable")
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:M{end}"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 13
    for col in "CDE":
        ws.column_dimensions[col].width = 23
    ws.column_dimensions["F"].width = 19
    for col in "GHI":
        ws.column_dimensions[col].width = 29
    for col in "JKLM":
        ws.column_dimensions[col].width = 25
    set_number_format(ws, "A6:A103", "0")
    set_number_format(ws, "C6:E103", "0.000E+00")
    set_number_format(ws, "F6:F103", "0.0")
    set_number_format(ws, "G6:M103", "0.000E+00")
    return ws


def build_pf_sheet(wb, best_rows):
    ws = wb.create_sheet("PF_over_tau")
    setup_sheet(ws)
    style_title(ws, "A1:H1", "Best sampled PF/tau values")
    ws.merge_cells("A2:H3")
    ws["A2"] = (
        "These are the best points within the discrete carrier-density grid, not a continuous optimum. "
        "PF/tau is used because no relaxation time has been assumed. The p/n comparison assumes the same tau."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].fill = PatternFill("solid", fgColor="FFF4E5")

    headers = [
        "Temperature (K)",
        "Type",
        "Density (cm^-3)",
        "S (uV/K)",
        "sigma/tau",
        "kappa_e/tau",
        "PF/tau (W m^-1 K^-2 s^-1)",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(5, col, header)
    style_header(ws["A5:G5"])

    for excel_row, row in enumerate(best_rows, start=6):
        ws.cell(excel_row, 1, int(float(row["temperature_K"])))
        ws.cell(excel_row, 2, row["carrier_type"])
        ws.cell(excel_row, 3, float(row["carrier_density_cm-3"]))
        ws.cell(excel_row, 4, float(row["seebeck_uV_K"]))
        ws.cell(excel_row, 5, float(row["sigma_over_tau_ohm-1_m-1_s-1"]))
        ws.cell(excel_row, 6, float(row["kappa_e_over_tau_W_m-1_K-1_s-1"]))
        ws.cell(excel_row, 7, f"=(D{excel_row}/1000000)^2*E{excel_row}")
    end = 5 + len(best_rows)
    add_table(ws, f"A5:G{end}", "BestPFTable")
    ws.freeze_panes = "A6"

    set_number_format(ws, "C6:C19", "0.00E+00")
    set_number_format(ws, "D6:D19", "0.0")
    set_number_format(ws, "E6:G19", "0.000E+00")
    widths = {"A": 17, "B": 10, "C": 21, "D": 14, "E": 18, "F": 20, "G": 29}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Visible, auditable helper table for the chart.
    helper_headers = ["Temperature (K)", "n-type PF/tau", "p-type PF/tau"]
    for col, header in enumerate(helper_headers, 10):
        ws.cell(22, col, header)
    style_header(ws["J22:L22"])
    temperatures = sorted({int(float(row["temperature_K"])) for row in best_rows})
    lookup = {(int(float(r["temperature_K"])), r["carrier_type"]): r for r in best_rows}
    for offset, temperature in enumerate(temperatures, start=23):
        ws.cell(offset, 10, temperature)
        for carrier, col in (("n", 11), ("p", 12)):
            row = lookup[(temperature, carrier)]
            s = float(row["seebeck_uV_K"])
            sigma = float(row["sigma_over_tau_ohm-1_m-1_s-1"])
            ws.cell(offset, col, (s / 1e6) ** 2 * sigma)
    set_number_format(ws, "K23:L29", "0.000E+00")
    ws.column_dimensions["J"].width = 17
    ws.column_dimensions["K"].width = 19
    ws.column_dimensions["L"].width = 19

    chart = LineChart()
    chart.title = "Best sampled PF/tau vs temperature"
    chart.style = 13
    chart.height = 8.2
    chart.width = 15.5
    chart.y_axis.title = "PF/tau (W m^-1 K^-2 s^-1)"
    chart.x_axis.title = "Temperature (K)"
    chart.y_axis.numFmt = "0.0E+00"
    chart.legend.position = "t"
    chart.add_data(Reference(ws, min_col=11, max_col=12, min_row=22, max_row=29), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=10, min_row=23, max_row=29))
    chart.series[0].graphicalProperties.line.solidFill = RED
    chart.series[0].graphicalProperties.line.width = 24000
    chart.series[0].marker.symbol = "circle"
    chart.series[0].marker.size = 6
    chart.series[1].graphicalProperties.line.solidFill = TEAL
    chart.series[1].graphicalProperties.line.width = 24000
    chart.series[1].marker.symbol = "circle"
    chart.series[1].marker.size = 6
    ws.add_chart(chart, "I5")

    ws.conditional_formatting.add(
        f"G6:G{end}",
        CellIsRule(operator="greaterThan", formula=["1E11"], fill=PatternFill("solid", fgColor=LIGHT_TEAL)),
    )


def build_convergence(wb, cutoff_rows, k_rows):
    ws = wb.create_sheet("Convergence")
    setup_sheet(ws)
    style_title(ws, "A1:N1", "Convergence checks")

    style_section(ws["A3"], "Plane-wave cutoff")
    ws.merge_cells("A3:E3")
    cutoff_headers = ["ecutwfc (Ry)", "ecutrho (Ry)", "Energy (Ry)", "Delta vs 100 Ry (meV/atom)", "Pressure (kbar)"]
    for col, header in enumerate(cutoff_headers, 1):
        ws.cell(4, col, header)
    style_header(ws["A4:E4"])
    for idx, row in enumerate(cutoff_rows, start=5):
        values = [
            int(float(row["ecutwfc_Ry"])),
            int(float(row["ecutrho_Ry"])),
            float(row["total_energy_Ry"]),
            float(row["delta_meV_per_atom_vs_max"]),
            float(row["pressure_kbar"]),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(idx, col, value)
    add_table(ws, f"A4:E{4 + len(cutoff_rows)}", "CutoffTable")
    set_number_format(ws, "C5:C10", "0.00000000")
    set_number_format(ws, "D5:E10", "0.000")

    cutoff_chart = LineChart()
    cutoff_chart.title = "Energy convergence with cutoff"
    cutoff_chart.height = 7
    cutoff_chart.width = 13
    cutoff_chart.y_axis.title = "Delta energy (meV/atom)"
    cutoff_chart.x_axis.title = "ecutwfc (Ry)"
    cutoff_chart.add_data(Reference(ws, min_col=4, min_row=4, max_row=10), titles_from_data=True)
    cutoff_chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=10))
    cutoff_chart.legend = None
    cutoff_chart.series[0].graphicalProperties.line.solidFill = TEAL
    cutoff_chart.series[0].marker.symbol = "circle"
    ws.add_chart(cutoff_chart, "G3")

    style_section(ws["A13"], "K-point mesh")
    ws.merge_cells("A13:E13")
    k_headers = ["Mesh", "Full k points", "Irreducible k points", "Energy (Ry)", "Delta vs 5x5x3 (meV/atom)"]
    for col, header in enumerate(k_headers, 1):
        ws.cell(14, col, header)
    style_header(ws["A14:E14"])
    for idx, row in enumerate(k_rows, start=15):
        values = [
            row["mesh"],
            int(float(row["full_kpoints"])),
            int(float(row["irreducible_kpoints"])),
            float(row["total_energy_Ry"]),
            float(row["delta_meV_per_atom_vs_max"]),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(idx, col, value)
    add_table(ws, f"A14:E{14 + len(k_rows)}", "KpointTable")
    set_number_format(ws, "D15:D18", "0.00000000")
    set_number_format(ws, "E15:E18", "0.000")

    k_chart = LineChart()
    k_chart.title = "Energy convergence with k-point mesh"
    k_chart.height = 7
    k_chart.width = 13
    k_chart.y_axis.title = "Delta energy (meV/atom)"
    k_chart.x_axis.title = "K-point mesh"
    k_chart.add_data(Reference(ws, min_col=5, min_row=14, max_row=18), titles_from_data=True)
    k_chart.set_categories(Reference(ws, min_col=1, min_row=15, max_row=18))
    k_chart.legend = None
    k_chart.series[0].graphicalProperties.line.solidFill = TEAL
    k_chart.series[0].marker.symbol = "circle"
    ws.add_chart(k_chart, "G13")

    for col, width in {"A": 17, "B": 18, "C": 23, "D": 30, "E": 30}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def build_charts(wb):
    ws = wb.create_sheet("Charts", 0)
    setup_sheet(ws)
    ws.sheet_view.zoomScale = 85
    style_title(ws, "A1:N1", "SrCu2SnS4 calculation plots")
    ws.merge_cells("A2:N2")
    ws["A2"] = (
        "PF/tau is shown because no relaxation time was assumed. The transport points are the best "
        "values within the sampled carrier-density grid; they are not continuous optima."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws.row_dimensions[2].height = 30

    pf_ws = wb["PF_over_tau"]
    pf_chart = LineChart()
    pf_chart.title = "Best sampled PF/tau vs temperature"
    pf_chart.style = 13
    pf_chart.height = 9
    pf_chart.width = 18
    pf_chart.y_axis.title = "PF/tau (W m^-1 K^-2 s^-1)"
    pf_chart.x_axis.title = "Temperature (K)"
    pf_chart.y_axis.numFmt = "0.0E+00"
    pf_chart.legend.position = "t"
    pf_chart.add_data(
        Reference(pf_ws, min_col=11, max_col=12, min_row=22, max_row=29),
        titles_from_data=True,
    )
    pf_chart.set_categories(Reference(pf_ws, min_col=10, min_row=23, max_row=29))
    pf_chart.series[0].graphicalProperties.line.solidFill = RED
    pf_chart.series[0].graphicalProperties.line.width = 24000
    pf_chart.series[0].marker.symbol = "circle"
    pf_chart.series[0].marker.size = 6
    pf_chart.series[1].graphicalProperties.line.solidFill = TEAL
    pf_chart.series[1].graphicalProperties.line.width = 24000
    pf_chart.series[1].marker.symbol = "circle"
    pf_chart.series[1].marker.size = 6
    ws.add_chart(pf_chart, "A4")

    convergence_ws = wb["Convergence"]
    cutoff_chart = LineChart()
    cutoff_chart.title = "Energy convergence with cutoff"
    cutoff_chart.height = 8
    cutoff_chart.width = 14
    cutoff_chart.y_axis.title = "Delta energy (meV/atom)"
    cutoff_chart.x_axis.title = "ecutwfc (Ry)"
    cutoff_chart.add_data(
        Reference(convergence_ws, min_col=4, min_row=4, max_row=10),
        titles_from_data=True,
    )
    cutoff_chart.set_categories(Reference(convergence_ws, min_col=1, min_row=5, max_row=10))
    cutoff_chart.legend = None
    cutoff_chart.series[0].graphicalProperties.line.solidFill = TEAL
    cutoff_chart.series[0].marker.symbol = "circle"
    ws.add_chart(cutoff_chart, "A22")

    k_chart = LineChart()
    k_chart.title = "Energy convergence with k-point mesh"
    k_chart.height = 8
    k_chart.width = 14
    k_chart.y_axis.title = "Delta energy (meV/atom)"
    k_chart.x_axis.title = "K-point mesh"
    k_chart.add_data(
        Reference(convergence_ws, min_col=5, min_row=14, max_row=18),
        titles_from_data=True,
    )
    k_chart.set_categories(Reference(convergence_ws, min_col=1, min_row=15, max_row=18))
    k_chart.legend = None
    k_chart.series[0].graphicalProperties.line.solidFill = RED
    k_chart.series[0].marker.symbol = "circle"
    ws.add_chart(k_chart, "H22")

    for col in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 12


def verify(path):
    wb = load_workbook(path, data_only=False)
    assert wb.sheetnames == ["Charts", "Notes", "Transport_Data", "PF_over_tau", "Convergence"]
    assert wb.active.title == "Charts"
    assert wb["Transport_Data"].max_row == 103
    assert wb["PF_over_tau"].max_row >= 29
    assert len(wb["Charts"]._charts) == 3
    assert len(wb["PF_over_tau"]._charts) == 1
    assert len(wb["Convergence"]._charts) == 2
    formulas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.value)
    assert formulas and all("#REF!" not in formula for formula in formulas)
    assert not any("electronic_zT" in str(cell.value) for ws in wb.worksheets for row in ws.iter_rows() for cell in row)
    print(f"Verified sheets: {', '.join(wb.sheetnames)}")
    print(f"Verified formulas: {len(formulas)}")
    print("Verified charts: 6 total; 3 visible on the Charts opening sheet")


def main():
    full_rows = read_csv(SOURCE / "results" / "transport_full.csv")
    best_rows = read_csv(SOURCE / "results" / "transport_best_power_factor.csv")
    cutoff_rows = read_csv(SOURCE / "qe" / "convergence" / "cutoff" / "cutoff_results.csv")
    k_rows = read_csv(SOURCE / "qe" / "convergence" / "kpoints" / "kpoint_results.csv")

    wb = Workbook()
    wb.properties.title = "My SrCu2SnS4 calculation results"
    wb.properties.subject = "QE and BoltzTraP2 calculation notes and data"
    wb.properties.creator = "Yuhan"
    wb.properties.keywords = "SrCu2SnS4, Quantum ESPRESSO, BoltzTraP2"
    wb.properties.description = "My first-pass calculation notes and data"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    build_notes(wb)
    build_transport_data(wb, full_rows)
    build_pf_sheet(wb, best_rows)
    build_convergence(wb, cutoff_rows, k_rows)
    build_charts(wb)
    wb.active = 0

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    ATTACHMENT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    shutil.copy2(OUTPUT, ATTACHMENT)
    verify(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Copied {ATTACHMENT}")


if __name__ == "__main__":
    main()
